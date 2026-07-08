"""Exportação formatada da planilha com o resultado da inclusão CADIN."""

import os
import shutil
import tempfile
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import config
from leitor_planilha import COLUNA_CPF_CNPJ_NORMALIZADO, COLUNA_DATA_NORMALIZADA, COLUNA_LINHA_ORIGINAL


COLUNAS_OCULTAS = {
    COLUNA_LINHA_ORIGINAL,
    COLUNA_CPF_CNPJ_NORMALIZADO,
    COLUNA_DATA_NORMALIZADA,
}


def preparar_exportacao(df):
    return df[[coluna for coluna in df.columns if coluna not in COLUNAS_OCULTAS]].copy()


def exportar_resultado(df, caminho_saida: str) -> None:
    pasta_saida = os.path.dirname(caminho_saida)
    if pasta_saida:
        os.makedirs(pasta_saida, exist_ok=True)

    export_df = preparar_exportacao(df)
    export_df.to_excel(caminho_saida, index=False, engine="openpyxl")
    formatar_excel(caminho_saida)


def caminho_checkpoint_csv(caminho_saida: str) -> str:
    base, _ = os.path.splitext(caminho_saida)
    return f"{base}_checkpoint.csv"


def caminho_checkpoint_local(caminho_saida: str) -> str:
    nome_base = Path(caminho_saida).stem
    pasta = Path(tempfile.gettempdir()) / "Inclusao_CADIN_checkpoints"
    return str(pasta / f"{nome_base}_checkpoint.csv")


def exportar_checkpoint_csv(df, caminho_saida: str, local: bool = False) -> str:
    caminho_csv = caminho_checkpoint_local(caminho_saida) if local else caminho_checkpoint_csv(caminho_saida)
    pasta_saida = os.path.dirname(caminho_csv)
    if pasta_saida:
        os.makedirs(pasta_saida, exist_ok=True)

    preparar_exportacao(df).to_csv(caminho_csv, index=False, sep=";", encoding="utf-8-sig")
    return caminho_csv


def copiar_checkpoint_para_resultados(caminho_saida: str, caminho_origem: Optional[str] = None) -> str:
    origem = caminho_origem or caminho_checkpoint_local(caminho_saida)
    destino = caminho_checkpoint_csv(caminho_saida)
    pasta_destino = os.path.dirname(destino)
    if pasta_destino:
        os.makedirs(pasta_destino, exist_ok=True)
    if os.path.exists(origem):
        shutil.copy2(origem, destino)
    return destino


def formatar_excel(caminho: str) -> None:
    wb = load_workbook(caminho)
    ws = wb.active
    ws.title = "Resultado CADIN"

    header_fill = PatternFill("solid", fgColor="12355B")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    status_colors = {
        config.STATUS_SUCESSO: "C6EFCE",
        config.STATUS_ERRO: "FFC7CE",
        config.STATUS_IGNORADO: "FFEB9C",
        config.STATUS_PARADO: "E7E6E6",
        config.STATUS_TESTE: "D9EAF7",
        config.STATUS_JA_INSCRITO: "FCE4D6",
    }
    headers = {cell.value: cell.column for cell in ws[1]}
    status_col = headers.get("STATUS_CADIN")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if status_col:
            status = str(ws.cell(row=row[0].row, column=status_col).value or "").upper()
            fill = status_colors.get(status)
            if fill:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill)

    for column_cells in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 60)

    if ws.max_row >= 2 and ws.max_column >= 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        tabela = Table(displayName="TabelaResultadoCADIN", ref=ref)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tabela)

    ws.freeze_panes = "A2"
    wb.save(caminho)

